import os
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, jsonify
from openpyxl import load_workbook

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "inventory.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "reports")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tool_type TEXT,
            serial_number TEXT,
            size TEXT,
            thread_type TEXT,
            location TEXT,
            status TEXT,
            report_link TEXT,
            description TEXT
        )
    ''')
    conn.commit()
    conn.close()

init_db()

# قالب‌فیلتر برای تبدیل ارقام فارسی/عربی به انگلیسی
@app.template_filter('en_digits')
def en_digits(value):
    if value is None:
        return ""
    s = str(value)
    mapping = {
        '۰':'0','۱':'1','۲':'2','۳':'3','۴':'4','۵':'5','۶':'6','۷':'7','۸':'8','۹':'9',
        '٠':'0','١':'1','٢':'2','٣':'3','٤':'4','٥':'5','٦':'6','٧':'7','٨':'8','٩':'9'
    }
    for k,v in mapping.items():
        s = s.replace(k, v)
    return s

@app.route("/", methods=["GET"])
def index():
    query = "SELECT * FROM inventory WHERE 1=1"
    params = []
    # filters: tool_type (select), serial_number (text), status (select), location (text)
    tool_type = request.args.get("tool_type", "").strip()
    serial_number = request.args.get("serial_number", "").strip()
    status = request.args.get("status", "").strip()
    location = request.args.get("location", "").strip()

    if tool_type:
        query += " AND tool_type LIKE ?"; params.append(f"%{tool_type}%")
    if serial_number:
        query += " AND serial_number LIKE ?"; params.append(f"%{serial_number}%")
    if status:
        query += " AND status LIKE ?"; params.append(f"%{status}%")
    if location:
        query += " AND location LIKE ?"; params.append(f"%{location}%")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(query, params)
    tools = c.fetchall()
    conn.close()
    filters = {"tool_type": tool_type, "serial_number": serial_number, "status": status, "location": location}
    return render_template("index.html", tools=tools, filters=filters)

@app.route("/add", methods=["POST"])
def add():
    tool_type = request.form.get("tool_type") or request.form.get("tool_type_select")
    serial_number = request.form.get("serial_number")
    size = request.form.get("size")
    thread_type = request.form.get("thread_type")
    location = request.form.get("location")
    status = request.form.get("status")
    description = request.form.get("description", "")
    report_file = request.files.get("report_file")
    report_link = ""
    if report_file and report_file.filename:
        save_path = os.path.join(UPLOAD_FOLDER, report_file.filename)
        report_file.save(save_path)
        report_link = "/static/reports/" + report_file.filename

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''INSERT INTO inventory
        (tool_type, serial_number, size, thread_type, location, status, report_link, description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
        (tool_type, serial_number, size, thread_type, location, status, report_link, description))
    conn.commit()
    conn.close()
    return redirect(url_for("index"))

@app.route("/edit/<int:id>", methods=["GET", "POST"])
def edit(id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    if request.method == "POST":
        tool_type = request.form.get("tool_type")
        serial_number = request.form.get("serial_number")
        size = request.form.get("size")
        thread_type = request.form.get("thread_type")
        location = request.form.get("location")
        status = request.form.get("status")
        description = request.form.get("description", "")
        c.execute('''UPDATE inventory SET
            tool_type=?, serial_number=?, size=?, thread_type=?, location=?, status=?, description=?
            WHERE id=?''',
            (tool_type, serial_number, size, thread_type, location, status, description, id))
        conn.commit()
        conn.close()
        return redirect(url_for("index"))
    c.execute("SELECT * FROM inventory WHERE id=?", (id,))
    item = c.fetchone()
    conn.close()
    return render_template("edit.html", item=item)

@app.route("/delete/<int:id>")
def delete(id):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM inventory WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for("index"))

@app.route("/upload_excel", methods=["POST"])
def upload_excel():
    excel_file = request.files.get("excel_file")
    if not excel_file or not excel_file.filename.endswith(".xlsx"):
        return "لطفاً فایل XLSX معتبر انتخاب کنید", 400
    wb = load_workbook(excel_file)
    sheet = wb.active
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        if len(row) < 6:
            continue
        tool_type, serial_number, size, thread_type, location, status = row[:6]
        c.execute('''INSERT INTO inventory
            (tool_type, serial_number, size, thread_type, location, status, report_link, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (tool_type, serial_number, size, thread_type, location, status, "", ""))
    conn.commit()
    conn.close()
    return redirect(url_for("index"))

@app.route("/update_description/<int:id>", methods=["POST"])
def update_description(id):
    if request.is_json:
        desc = request.get_json().get("description", "")
    else:
        desc = request.form.get("description", "")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE inventory SET description=? WHERE id=?", (desc, id))
    conn.commit()
    conn.close()
    if request.is_json:
        return jsonify({"success": True})
    ref = request.referrer
    return redirect(ref or url_for("index"))

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
